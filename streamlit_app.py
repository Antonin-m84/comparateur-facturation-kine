"""
🏥 Comparateur Facturation Kiné vs Rapport Hôpital
Application Streamlit pour comparer les facturations

Auteur: Antonin
"""

import streamlit as st
import pandas as pd
import re
from datetime import datetime
from io import BytesIO
import tempfile
import os
import traceback

# Patch pour corriger le bug 'biltinId' dans certains fichiers Excel
# Certains fichiers Excel mal formés contiennent 'biltinId' au lieu de 'builtinId'
try:
    from openpyxl.styles.named_styles import _NamedCellStyle
    _original_named_cell_style_init = _NamedCellStyle.__init__
    def _patched_named_cell_style_init(self, *args, **kwargs):
        if 'biltinId' in kwargs:
            kwargs['builtinId'] = kwargs.pop('biltinId')
        return _original_named_cell_style_init(self, *args, **kwargs)
    _NamedCellStyle.__init__ = _patched_named_cell_style_init
except Exception:
    pass

try:
    from openpyxl.styles.named_styles import NamedStyle
    _original_named_style_init = NamedStyle.__init__
    def _patched_named_style_init(self, *args, **kwargs):
        if 'biltinId' in kwargs:
            kwargs['builtinId'] = kwargs.pop('biltinId')
        return _original_named_style_init(self, *args, **kwargs)
    NamedStyle.__init__ = _patched_named_style_init
except Exception:
    pass

# Configuration de la page
st.set_page_config(
    page_title="Comparateur Facturation Kiné",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personnalisé pour une meilleure apparence
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1E88E5;
        text-align: center;
        margin-bottom: 2rem;
    }
    .success-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #E8F5E9;
        border-left: 5px solid #4CAF50;
        margin: 1rem 0;
    }
    .warning-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #FFF3E0;
        border-left: 5px solid #FF9800;
        margin: 1rem 0;
    }
    .error-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #FFEBEE;
        border-left: 5px solid #F44336;
        margin: 1rem 0;
    }
    .stDownloadButton > button {
        width: 100%;
        background-color: #4CAF50;
        color: white;
    }
</style>
""", unsafe_allow_html=True)


# =============================================================================
# FONCTIONS DE TRAITEMENT
# =============================================================================

def parse_billing_code(code_str):
    """Parse les codes de facturation (ex: 'K-1 + RECOND' -> ['K-1', 'RECOND'])"""
    codes = []
    code_str = code_str.strip().upper()
    
    if 'M24' in code_str:
        codes.append('M 24')
    if 'M6' in code_str or 'M 6' in code_str:
        codes.append('M 6')
    if 'K-1' in code_str:
        codes.append('K-1')
    if 'RECOND' in code_str:
        codes.append('RECOND')
    if 'K3/4' in code_str:
        codes.append('K3/4')
    if 'K20' in code_str:
        codes.append('K 20')
    if 'K15' in code_str:
        codes.append('K 15')
    
    return codes


def normalize_code(code):
    """Normalise un code pour comparaison"""
    code = code.strip().upper()
    code = re.sub(r'\s+', ' ', code)
    return code


def normalize_name(name):
    """Normalise un nom pour comparaison"""
    name = name.strip().upper()
    name = re.sub(r'^\([A-Z]\)\s*', '', name)
    name = name.replace(',', '')
    name = re.sub(r'\s+', ' ', name)
    return name


def read_excel_robust(file_bytes, sheet_name=0):
    """
    Lit un fichier Excel de manière robuste, en gérant les fichiers protégés
    et les erreurs de styles corrompus (biltinId, etc.)
    """
    
    # Méthode 1: Lecture standard avec pandas
    try:
        return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None, engine='openpyxl')
    except Exception as e1:
        pass
    
    # Méthode 2: Utiliser openpyxl avec data_only pour ignorer les formules
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]
        data = list(ws.values)
        wb.close()
        return pd.DataFrame(data)
    except Exception as e2:
        pass
    
    # Méthode 3: Sauvegarder temporairement et réouvrir (simule l'activation de modification)
    try:
        from openpyxl import load_workbook
        
        # Créer un fichier temporaire
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        
        # Ouvrir et resauvegarder pour "activer la modification"
        wb = load_workbook(tmp_path)
        
        # Récupérer les données
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        
        wb.close()
        os.unlink(tmp_path)  # Supprimer le fichier temporaire
        
        return pd.DataFrame(data)
    except Exception as e3:
        pass
    
    # Méthode 4: Lecture avec xlrd pour les anciens fichiers .xls
    try:
        return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None, engine='xlrd')
    except Exception as e4:
        pass
    
    # Méthode 5: Dernière tentative avec calamine (si disponible)
    try:
        return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None, engine='calamine')
    except Exception as e5:
        pass
    
    # Si toutes les méthodes échouent
    raise Exception(
        "Impossible de lire le fichier Excel. Le fichier est peut-être protégé ou corrompu.\n"
        "💡 Solution: Ouvrez le fichier dans Excel, cliquez sur 'Activer la modification', "
        "puis sauvegardez-le avant de le réuploader."
    )


def get_sheet_names_robust(file_bytes):
    """Récupère les noms des feuilles d'un fichier Excel"""
    errors = []
    
    # Méthode 1: pandas avec openpyxl
    try:
        xl = pd.ExcelFile(BytesIO(file_bytes), engine='openpyxl')
        sheets = xl.sheet_names
        xl.close()
        return sheets
    except Exception as e1:
        errors.append(f"Méthode pandas/openpyxl: {str(e1)}")
    
    # Méthode 2: openpyxl direct read_only
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e2:
        errors.append(f"Méthode openpyxl read_only: {str(e2)}")
    
    # Méthode 3: openpyxl sans read_only
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e3:
        errors.append(f"Méthode openpyxl standard: {str(e3)}")
    
    # Méthode 4: fichier temporaire
    try:
        from openpyxl import load_workbook
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        
        wb = load_workbook(tmp_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        os.unlink(tmp_path)
        return sheets
    except Exception as e4:
        errors.append(f"Méthode fichier temporaire: {str(e4)}")
    
    # Méthode 5: xlrd pour les anciens fichiers .xls
    try:
        xl = pd.ExcelFile(BytesIO(file_bytes), engine='xlrd')
        sheets = xl.sheet_names
        xl.close()
        return sheets
    except Exception as e5:
        errors.append(f"Méthode xlrd: {str(e5)}")
    
    # Si toutes les méthodes ont échoué, lever une exception avec toutes les erreurs
    error_details = "\n".join(errors)
    raise Exception(f"Impossible de lire les feuilles du fichier Excel.\nDétails des erreurs:\n{error_details}")


def parse_my_billing(file_bytes, sheet_name=0):
    """Parse le fichier de facturation personnelle"""
    df = read_excel_robust(file_bytes, sheet_name=sheet_name)
    
    records = []
    current_date = None
    
    for idx, row in df.iterrows():
        date_val = row.iloc[0] if pd.notna(row.iloc[0]) else None
        
        if date_val:
            if hasattr(date_val, 'day'):
                current_date = date_val.day
            elif isinstance(date_val, str):
                match = re.search(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', date_val)
                if match:
                    current_date = int(match.group(1))
        
        # Colonnes: A=Date, B=Dossier, C=Nom, D=Code
        dossier = row.iloc[1] if len(row) > 1 and pd.notna(row.iloc[1]) else None
        nom = row.iloc[2] if len(row) > 2 and pd.notna(row.iloc[2]) else None
        codes_str = row.iloc[3] if len(row) > 3 and pd.notna(row.iloc[3]) else None
        
        if dossier and nom and codes_str and current_date:
            codes = parse_billing_code(str(codes_str))
            for code in codes:
                records.append({
                    'date': current_date,
                    'dossier': str(int(dossier)) if isinstance(dossier, float) else str(dossier),
                    'nom': normalize_name(str(nom)),
                    'code': normalize_code(code),
                    'source': 'MA_FACTURATION'
                })
    
    return records


def parse_hospital_report(file_bytes, sheet_name=0):
    """Parse le rapport de l'hôpital"""
    df = read_excel_robust(file_bytes, sheet_name=sheet_name)
    
    records = []
    current_name = None
    current_dossier = None
    
    for idx, row in df.iterrows():
        nom_val = row.iloc[0] if pd.notna(row.iloc[0]) else None
        dossier_val = row.iloc[1] if len(row) > 1 and pd.notna(row.iloc[1]) else None
        code_val = row.iloc[2] if len(row) > 2 and pd.notna(row.iloc[2]) else None
        
        # Ignorer les en-têtes
        if nom_val and str(nom_val).strip().upper() in ['PATIENT', 'NOM', 'NOM PATIENT']:
            continue
        if dossier_val and str(dossier_val).strip().upper() in ['DOSSIER', 'N° DOSSIER', 'NUMERO DOSSIER', 'N°DOSSIER']:
            continue
        if code_val and str(code_val).strip().upper() in ['CODE', 'CODE INTERNE', 'CODES']:
            continue
        
        if nom_val:
            current_name = normalize_name(str(nom_val))
            current_dossier = str(int(dossier_val)) if dossier_val and isinstance(dossier_val, float) else str(dossier_val) if dossier_val else None
        
        if current_dossier and code_val:
            code = normalize_code(str(code_val))
            
            for day in range(1, 32):
                col_idx = 6 + (day - 1)
                if col_idx < len(row):
                    cell_val = row.iloc[col_idx]
                    if pd.notna(cell_val) and str(cell_val).strip() != '':
                        records.append({
                            'date': day,
                            'dossier': current_dossier,
                            'nom': current_name,
                            'code': code,
                            'source': 'RAPPORT_HOPITAL'
                        })
    
    return records


def log_error(error_message, error_traceback):
    """Enregistre l'erreur dans un fichier log"""
    log_dir = "error_logs"
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f"{log_dir}/error_{timestamp}.txt"
    
    with open(log_filename, 'w', encoding='utf-8') as f:
        f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"="*50 + "\n")
        f.write(f"Message d'erreur:\n{error_message}\n\n")
        f.write(f"Traceback complet:\n{error_traceback}\n")
    
    return log_filename


def compare_records(my_records, hospital_records):
    """Compare les enregistrements entre les deux sources"""
    
    my_filtered = my_records
    hosp_filtered = hospital_records
    
    def make_key(r):
        return (r['dossier'], r['code'], r['date'])
    
    my_keys = {make_key(r): r for r in my_filtered}
    hosp_keys = {make_key(r): r for r in hosp_filtered}
    
    only_in_my = []
    only_in_hospital = []
    matched = []
    
    for key, record in my_keys.items():
        if key in hosp_keys:
            matched.append(record)
        else:
            only_in_my.append(record)
    
    for key, record in hosp_keys.items():
        if key not in my_keys:
            only_in_hospital.append(record)
    
    return matched, only_in_my, only_in_hospital


def create_excel_output(my_records, hospital_records, only_mine, only_hospital, sheet_name):
    """Crée le fichier Excel de sortie"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Feuille 1: Différences
        differences = []
        for r in only_mine:
            diff = r.copy()
            diff['statut'] = '⚠️ Manquant dans rapport hôpital'
            differences.append(diff)
        for r in only_hospital:
            diff = r.copy()
            diff['statut'] = '⚠️ Manquant dans ma facturation'
            differences.append(diff)
        
        if differences:
            df_diff = pd.DataFrame(differences)
            df_diff = df_diff.sort_values(['date', 'dossier', 'statut']).reset_index(drop=True)
            df_diff.to_excel(writer, sheet_name='Differences', index=False)
        else:
            df_success = pd.DataFrame([{'Message': '✅ Bravo tout est en ordre !'}])
            df_success.to_excel(writer, sheet_name='Differences', index=False)
        
        # Feuille 2: Ma facturation
        if my_records:
            df_my = pd.DataFrame(my_records)
            df_my = df_my.sort_values(['date', 'dossier']).reset_index(drop=True)
            df_my.to_excel(writer, sheet_name='Ma_Facturation', index=False)
        
        # Feuille 3: Rapport hôpital
        if hospital_records:
            df_hosp = pd.DataFrame(hospital_records)
            df_hosp = df_hosp.sort_values(['date', 'dossier']).reset_index(drop=True)
            df_hosp.to_excel(writer, sheet_name='Rapport_Hopital', index=False)
    
    output.seek(0)
    return output


# =============================================================================
# INTERFACE STREAMLIT
# =============================================================================

def main():
    # En-tête
    st.markdown('<h1 class="main-header">🏥 Comparateur Facturation Kiné</h1>', unsafe_allow_html=True)
    st.markdown("---")
    
    # Sidebar avec instructions
    with st.sidebar:
        st.header("📖 Mode d'emploi")
        st.markdown("""
        **1.** Uploadez votre fichier de facturation
        
        **2.** Uploadez le rapport de l'hôpital
        
        **3.** Sélectionnez les feuilles Excel (mois)
        
        **4.** Cliquez sur **Lancer la comparaison**
        
        **5.** Téléchargez le fichier résultat
        """)
        
        st.markdown("---")
        st.header("⚠️ En cas d'erreur")
        st.info("""
        Si vous avez une erreur avec le fichier de l'hôpital:
        
        1. Ouvrez le fichier dans Excel
        2. Cliquez sur **"Activer la modification"**
        3. Sauvegardez le fichier (Ctrl+S)
        4. Réuploadez le fichier ici
        """)
        
        st.markdown("---")
        st.caption("v1.0 - Développé pour les kinés par Antonin")
    
    # Colonnes principales
    col1, col2 = st.columns(2)
    
    # Upload fichier facturation
    with col1:
        st.subheader("📁 Votre fichier de facturation")
        uploaded_billing = st.file_uploader(
            "Glissez votre fichier Excel ici",
            type=['xlsx', 'xls'],
            key="billing",
            help="Créer une copie datée de votre fichier de facturation personnelle en local sur votre ordinateur"
        )
        
        billing_bytes = None
        selected_billing_sheet = None
        billing_sheets = ["Feuille1"]
        
        if uploaded_billing:
            try:
                billing_bytes = uploaded_billing.read()
                try:
                    billing_sheets = get_sheet_names_robust(billing_bytes)
                    st.info(f"📋 Feuilles trouvées: {', '.join(billing_sheets)}")
                except Exception as sheet_err:
                    error_tb = traceback.format_exc()
                    log_file = log_error(str(sheet_err), error_tb)
                    st.warning(f"⚠️ Impossible de lire les feuilles: {sheet_err}")
                    st.caption(f"📝 Erreur enregistrée dans: {log_file}")
                    with st.expander("🔍 Voir le détail de l'erreur"):
                        st.code(error_tb)
                    billing_sheets = ["Feuille1"]
                
                selected_billing_sheet = st.selectbox(
                    "Choisissez la feuille:",
                    billing_sheets,
                    key="billing_sheet"
                )
                st.success(f"✅ Fichier chargé: {uploaded_billing.name}")
            except Exception as e:
                error_tb = traceback.format_exc()
                log_file = log_error(str(e), error_tb)
                st.error(f"❌ Erreur: {e}")
                st.caption(f"📝 Erreur enregistrée dans: {log_file}")
                with st.expander("🔍 Voir le détail de l'erreur"):
                    st.code(error_tb)
                billing_bytes = None
    
    # Upload fichier hôpital
    with col2:
        st.subheader("🏥 WebRapport du GHDC version Excel")
        uploaded_hospital = st.file_uploader(
            "Glissez le rapport Excel ici",
            type=['xlsx', 'xls'],
            key="hospital",
            help="Le rapport de facturation de l'hôpital"
        )
        
        hospital_bytes = None
        selected_hospital_sheet = None
        hospital_sheets = ["Feuille1"]
        
        if uploaded_hospital:
            try:
                hospital_bytes = uploaded_hospital.read()
                try:
                    hospital_sheets = get_sheet_names_robust(hospital_bytes)
                    st.info(f"📋 Feuilles trouvées: {', '.join(hospital_sheets)}")
                except Exception as sheet_err:
                    error_tb = traceback.format_exc()
                    log_file = log_error(str(sheet_err), error_tb)
                    st.warning(f"⚠️ Impossible de lire les feuilles: {sheet_err}")
                    st.caption(f"📝 Erreur enregistrée dans: {log_file}")
                    with st.expander("🔍 Voir le détail de l'erreur"):
                        st.code(error_tb)
                    hospital_sheets = ["Feuille1"]
                
                selected_hospital_sheet = st.selectbox(
                    "Choisissez la feuille:",
                    hospital_sheets,
                    key="hospital_sheet"
                )
                st.success(f"✅ Fichier chargé: {uploaded_hospital.name}")
            except Exception as e:
                error_tb = traceback.format_exc()
                log_file = log_error(str(e), error_tb)
                st.error(f"❌ Erreur: {e}")
                st.caption(f"📝 Erreur enregistrée dans: {log_file}")
                with st.expander("🔍 Voir le détail de l'erreur"):
                    st.code(error_tb)
                st.warning("💡 Essayez d'ouvrir le fichier dans Excel, cliquez sur 'Activer la modification', sauvegardez et réuploadez.")
                hospital_bytes = None
    
    st.markdown("---")
    
    # Bouton de lancement
    if st.button("🚀 Lancer la comparaison", type="primary", use_container_width=True):
        
        if not billing_bytes:
            st.error("❌ Veuillez d'abord uploader votre fichier de facturation")
            return
        
        if not hospital_bytes:
            st.error("❌ Veuillez d'abord uploader le rapport de l'hôpital")
            return
        
        # Barre de progression
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        try:
            # Parsing facturation
            status_text.text("📊 Lecture de votre facturation...")
            progress_bar.progress(20)
            my_records = parse_my_billing(billing_bytes, sheet_name=selected_billing_sheet)
            
            # Parsing hôpital
            status_text.text("🏥 Lecture du rapport hôpital...")
            progress_bar.progress(50)
            hospital_records = parse_hospital_report(hospital_bytes, sheet_name=selected_hospital_sheet)
            
            # Comparaison
            status_text.text("🔍 Comparaison en cours...")
            progress_bar.progress(80)
            matched, only_mine, only_hospital = compare_records(
                my_records, hospital_records
            )
            
            progress_bar.progress(100)
            status_text.text("✅ Comparaison terminée!")
            
            # Affichage des résultats
            st.markdown("---")
            st.header("📊 Résultats")
            
            # Métriques
            col_m1, col_m2, col_m3 = st.columns(3)
            with col_m1:
                st.metric("✅ Appariées", len(matched))
            with col_m2:
                st.metric("⚠️ Manquantes hôpital", len(only_mine))
            with col_m3:
                st.metric("⚠️ Manquantes chez vous", len(only_hospital))
            
            # Détails des différences
            if only_mine:
                st.subheader("⚠️ Dans votre facturation mais PAS dans le rapport hôpital")
                df_only_mine = pd.DataFrame(only_mine)
                df_only_mine = df_only_mine.sort_values(['date', 'dossier'])
                st.dataframe(df_only_mine, use_container_width=True)
            
            if only_hospital:
                st.subheader("⚠️ Dans le rapport hôpital mais PAS dans votre facturation")
                df_only_hospital = pd.DataFrame(only_hospital)
                df_only_hospital = df_only_hospital.sort_values(['date', 'dossier'])
                st.dataframe(df_only_hospital, use_container_width=True)
            
            if not only_mine and not only_hospital:
                st.balloons()
                st.success("🎉 **Bravo ! Tout est en ordre, aucune différence trouvée !**")
            
            # Téléchargement
            st.markdown("---")
            st.subheader("📥 Télécharger le rapport complet")
            
            timestamp = datetime.now().strftime("%d%m%Y_%H%M")
            safe_sheet_name = re.sub(r'[\\/*?:"<>|]', '_', selected_billing_sheet or "export")
            output_filename = f"comparaison_{safe_sheet_name}_{timestamp}.xlsx"
            
            excel_output = create_excel_output(
                my_records, hospital_records, only_mine, only_hospital, selected_billing_sheet
            )
            
            st.download_button(
                label="📥 Télécharger le fichier Excel",
                data=excel_output,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
        except Exception as e:
            progress_bar.empty()
            status_text.empty()
            error_tb = traceback.format_exc()
            log_file = log_error(str(e), error_tb)
            st.error(f"❌ Une erreur s'est produite: {e}")
            st.caption(f"📝 Erreur enregistrée dans: {log_file}")
            with st.expander("🔍 Voir le détail de l'erreur"):
                st.code(error_tb)
            st.warning("""
            💡 **Solutions possibles:**
            - Vérifiez que vos fichiers sont au bon format
            - Pour le rapport hôpital: ouvrez-le dans Excel, cliquez sur "Activer la modification", sauvegardez et réuploadez
            - Vérifiez que vous avez sélectionné les bonnes feuilles
            """)


if __name__ == "__main__":
    main()
